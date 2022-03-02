from openpyxl import load_workbook
import numpy as np
from pathlib import Path

#create setup arrays
file_names = ['']*12
sheet_names = ['Data - Cars','Data - HT','Data - Pedestrians (Crosswalks)']
number_of_files = 12
sheets_per_file = 4
most_rows = 16
most_cols = 16
route_decisions = np.zeros((number_of_files,sheets_per_file,most_rows,most_cols),int)
i = 0

direc_str = "E:\\Graduate\\Simulation\\Stantec_Intersections" #Location of files of interest

#check if directory has files
try:
    #breaks if directory is empty
    next(Path(direc_str).iterdir(), None)
except:
    print('\nChange Directory\n')
    exit()

#grab all files ending in .xlsx
pathlist = Path(direc_str).glob('**/*.xlsx')
for path in pathlist:
    # because path is object not string
    file_names[i] = str(path)

    wb = load_workbook(file_names[i])
    j = 0 
    #iterate over sheets
    for sheet in sheet_names:
        #activate worksheet
        ws = wb[sheet_names[j]]
        k = 0
        #collect car/HT data
        if sheet_names[j] != 'Data - Pedestrians (Crosswalks)':
            for row in ws.iter_rows(min_row=37,min_col=2, max_row=44, max_col=17, values_only=True):
                #place row from Excel into array row
                route_decisions[i][j][k] = row
                k = k + 1
            for row in ws.iter_rows(min_row=289,min_col=2, max_row=296, max_col=17, values_only=True):
                route_decisions[i][j][k] = row
                k = k + 1
        #collect ped data
        else:
            #for non-diagonal-crossing intersections
            if ws[7][9].value == 'TOTAL':
                for row in ws.iter_rows(min_row=36,min_col=2, max_row=43, max_col=9, values_only=True):
                    route_decisions[i][j][k] = np.concatenate((row,[0,0,0,0,0,0,0,0]),axis=None)
                    k = k + 1
                for row in ws.iter_rows(min_row=286,min_col=2, max_row=293, max_col=9, values_only=True):
                    route_decisions[i][j][k] = np.concatenate((row,[0,0,0,0,0,0,0,0]),axis=None)
                    k = k + 1
            #for diagonal-crossing intersections
            else:
                for row in ws.iter_rows(min_row=36,min_col=2, max_row=43, max_col=13, values_only=True):
                    route_decisions[i][j][k] = np.concatenate((row,[0,0,0,0]),axis=None)
                    k = k + 1
                for row in ws.iter_rows(min_row=286,min_col=2, max_row=293, max_col=13, values_only=True):
                    route_decisions[i][j][k] = np.concatenate((row,[0,0,0,0]),axis=None)
                    k = k + 1
        j = j + 1
    i = i + 1

i = 0
j = 0
k = 0
mwf_files = np.zeros(6,int)
print(mwf_files)
tth_files = np.zeros(6,int)
for i in range(len(file_names)):
    if file_names[i].find('Monday') != -1 or file_names[i].find('Wednesday') != -1 or file_names[i].find('Friday') != -1:
        mwf_files[j] = i
        j = j + 1
    elif file_names[i].find('Tuesday') != -1 or file_names[i].find('Thursday') != -1:
        tth_files[k] = i
        k = k + 1

print(mwf_files)
print(tth_files)